import openpyxl
import datetime
from datetime import timedelta
from calendar import mdays
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string
from openpyxl.styles import Font


def header():
	italic24Font = Font(size=18, italic=False)
	wb=load_workbook('testing.xlsx',read_only = False, data_only = True)
	sheet =wb.get_sheet_by_name('Sheet')
	wb.active
	month="{:%B %Y}".format(datetime.date.today())
	today = datetime.date.today()
	sheet['a1'].font=italic24Font
	sheet['b1'].font=italic24Font
	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 20
	sheet['a1']="UID"
	sheet['b1']="Name"
	start_col =3
	start_month=get_column_letter(start_col)
	sheet['%s1' % start_month] = "August 2016"
	
	curr_col = sheet.max_column
	#print (curr_col)
	this_month=get_column_letter(curr_col)
	if sheet.cell('%s1' % this_month).value != month:
		sheet['%s1' % this_month] = month
	
	sheet.column_dimensions['%s' % this_month].width = 20
	wb.save('testing.xlsx')


